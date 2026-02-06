from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func, and_, or_
from datetime import datetime
from io import BytesIO
from typing import Optional, List

from app.db.session import get_db
from app.models import AccessRequest, User, Approval, System, Subsystem, AccessRole
from app.models.request import RequestStatus
from app.api.deps import get_current_superuser

router = APIRouter()


def get_status_label(status_value):
    """Получить русское название статуса"""
    labels = {
        'draft': 'Черновик',
        'submitted': 'Отправлена',
        'in_review': 'На рассмотрении',
        'approved': 'Одобрена',
        'rejected': 'Отклонена',
        'implemented': 'Выполнена',
        'cancelled': 'Отменена',
        'expired': 'Истекла',
    }
    return labels.get(status_value, status_value)


def get_type_label(type_value):
    """Получить русское название типа заявки"""
    labels = {
        'new_access': 'Новый доступ',
        'modify_access': 'Изменение доступа',
        'revoke_access': 'Отзыв доступа',
        'temporary_access': 'Временный доступ',
    }
    return labels.get(type_value, type_value)


def get_approval_status_label(status_value):
    """Получить русское название статуса согласования"""
    labels = {
        'pending': 'Ожидает',
        'approved': 'Одобрено',
        'rejected': 'Отклонено',
    }
    return labels.get(status_value, status_value)


def format_date(dt):
    """Форматировать дату"""
    if not dt:
        return '—'
    if isinstance(dt, str):
        return dt[:10]
    return dt.strftime('%d.%m.%Y %H:%M')


def get_all_requests(db: Session):
    """Получить все заявки с полной информацией"""
    requests = db.query(AccessRequest).options(
        joinedload(AccessRequest.requester),
        joinedload(AccessRequest.target_user),
        joinedload(AccessRequest.system),
        joinedload(AccessRequest.subsystem),
        joinedload(AccessRequest.access_role),
        joinedload(AccessRequest.approvals).joinedload(Approval.approver),
        joinedload(AccessRequest.comments),
    ).order_by(AccessRequest.created_at.desc()).all()
    return requests


@router.get("/excel")
async def export_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_superuser)
):
    """Экспорт всех заявок в Excel (только для администраторов)"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    requests = get_all_requests(db)

    wb = Workbook()

    # Стили
    header_fill = PatternFill(start_color="16306C", end_color="16306C", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    data_font = Font(size=10)
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # ===== Лист 1: Все заявки =====
    ws = wb.active
    ws.title = "Заявки"

    headers = [
        'ID', 'Номер заявки', 'Статус', 'Тип заявки', 'Система', 'Подсистема',
        'Роль доступа', 'Для сотрудника', 'Email получателя', 'Инициатор', 'Email инициатора',
        'Дата создания', 'Дата отправки', 'Дата завершения', 'Текущий этап',
        'Временный доступ', 'Действителен с', 'Действителен до', 'Цель / Обоснование'
    ]

    # Заголовки
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_alignment

    # Данные
    for row_idx, req in enumerate(requests, 2):
        data = [
            req.id,
            req.request_number,
            get_status_label(req.status.value if req.status else '—'),
            get_type_label(req.request_type.value if req.request_type else '—'),
            req.system.name if req.system else '—',
            req.subsystem.name if req.subsystem else '—',
            req.access_role.name if req.access_role else '—',
            req.target_user.full_name if req.target_user else '—',
            req.target_user.email if req.target_user else '—',
            req.requester.full_name if req.requester else '—',
            req.requester.email if req.requester else '—',
            format_date(req.created_at),
            format_date(req.submitted_at),
            format_date(req.completed_at),
            req.current_step,
            'Да' if req.is_temporary else 'Нет',
            format_date(req.valid_from) if req.is_temporary else '—',
            format_date(req.valid_until) if req.is_temporary else '—',
            req.purpose or '—',
        ]

        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = left_alignment if col_idx > 1 else center_alignment

    # Ширина колонок
    column_widths = [6, 18, 15, 18, 20, 15, 18, 22, 25, 22, 25, 16, 16, 16, 12, 15, 16, 16, 40]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Закрепить заголовок
    ws.freeze_panes = 'A2'

    # ===== Лист 2: Согласования =====
    ws2 = wb.create_sheet("Согласования")

    approval_headers = [
        'ID заявки', 'Номер заявки', 'Этап', 'Согласующий', 'Email', 'Роль',
        'Статус', 'Дата решения', 'Комментарий'
    ]

    for col, header in enumerate(approval_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_alignment

    row_idx = 2
    for req in requests:
        for approval in sorted(req.approvals, key=lambda a: a.step_number):
            data = [
                req.id,
                req.request_number,
                approval.step_number,
                approval.approver.full_name if approval.approver else '—',
                approval.approver.email if approval.approver else '—',
                approval.approver_role or '—',
                get_approval_status_label(approval.status.value if approval.status else '—'),
                format_date(approval.decision_date),
                approval.comment or '—',
            ]

            for col_idx, value in enumerate(data, 1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = left_alignment if col_idx > 2 else center_alignment

            row_idx += 1

    approval_widths = [10, 18, 8, 25, 30, 20, 15, 18, 40]
    for col, width in enumerate(approval_widths, 1):
        ws2.column_dimensions[get_column_letter(col)].width = width

    ws2.freeze_panes = 'A2'

    # ===== Лист 3: Статистика =====
    ws3 = wb.create_sheet("Статистика")

    # Заголовок
    ws3.cell(row=1, column=1, value="Статистика по заявкам").font = Font(bold=True, size=14)
    ws3.merge_cells('A1:B1')

    stats = [
        ('', ''),
        ('Всего заявок', len(requests)),
        ('Черновики', sum(1 for r in requests if r.status and r.status.value == 'draft')),
        ('На рассмотрении', sum(1 for r in requests if r.status and r.status.value == 'in_review')),
        ('Одобрено', sum(1 for r in requests if r.status and r.status.value == 'approved')),
        ('Отклонено', sum(1 for r in requests if r.status and r.status.value == 'rejected')),
        ('Выполнено', sum(1 for r in requests if r.status and r.status.value == 'implemented')),
        ('Отменено', sum(1 for r in requests if r.status and r.status.value == 'cancelled')),
        ('', ''),
        ('Дата выгрузки', datetime.now().strftime('%d.%m.%Y %H:%M')),
    ]

    for row_idx, (label, value) in enumerate(stats, 2):
        if label:
            ws3.cell(row=row_idx, column=1, value=label).font = Font(bold=True, size=11)
            ws3.cell(row=row_idx, column=2, value=value).font = Font(size=11)
            ws3.cell(row=row_idx, column=1).border = thin_border
            ws3.cell(row=row_idx, column=2).border = thin_border

    ws3.column_dimensions['A'].width = 25
    ws3.column_dimensions['B'].width = 20

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"zajavki_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/user-access")
async def get_user_access_report(
    system_id: Optional[int] = None,
    department: Optional[str] = None,
    search: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_superuser)
):
    """
    Отчёт по пользователям - кто какие доступы имеет в каких системах.
    Показывает только одобренные и выполненные заявки (реальные доступы).
    """
    # Базовый запрос - заявки со статусом approved или implemented
    query = db.query(AccessRequest).options(
        joinedload(AccessRequest.target_user),
        joinedload(AccessRequest.system),
        joinedload(AccessRequest.subsystem),
        joinedload(AccessRequest.access_role),
    ).filter(
        AccessRequest.status.in_([RequestStatus.APPROVED, RequestStatus.IMPLEMENTED])
    )

    # Фильтр по системе
    if system_id:
        query = query.filter(AccessRequest.system_id == system_id)

    # Фильтр по отделу
    if department:
        query = query.join(AccessRequest.target_user).filter(
            User.department.ilike(f"%{department}%")
        )

    # Поиск по имени пользователя
    if search:
        if not department:  # Если join ещё не сделан
            query = query.join(AccessRequest.target_user)
        query = query.filter(
            or_(
                User.full_name.ilike(f"%{search}%"),
                User.username.ilike(f"%{search}%"),
                User.email.ilike(f"%{search}%")
            )
        )

    # Общее количество
    total = query.count()

    # Получить данные с пагинацией
    requests = query.order_by(
        AccessRequest.target_user_id,
        AccessRequest.system_id
    ).offset(skip).limit(limit).all()

    # Группировка по пользователям
    users_access = {}
    for req in requests:
        user_id = req.target_user_id
        if user_id not in users_access:
            users_access[user_id] = {
                "user_id": user_id,
                "full_name": req.target_user.full_name if req.target_user else "—",
                "username": req.target_user.username if req.target_user else "—",
                "email": req.target_user.email if req.target_user else "—",
                "department": req.target_user.department if req.target_user else "—",
                "position": req.target_user.position if req.target_user else "—",
                "systems": []
            }

        users_access[user_id]["systems"].append({
            "request_id": req.id,
            "request_number": req.request_number,
            "system_id": req.system_id,
            "system_name": req.system.name if req.system else "—",
            "system_code": req.system.code if req.system else "—",
            "subsystem_name": req.subsystem.name if req.subsystem else None,
            "access_role": req.access_role.name if req.access_role else "—",
            "access_level": req.access_role.access_level if req.access_role else "—",
            "status": req.status.value,
            "is_temporary": req.is_temporary,
            "valid_from": req.valid_from.isoformat() if req.valid_from else None,
            "valid_until": req.valid_until.isoformat() if req.valid_until else None,
            "granted_at": req.completed_at.isoformat() if req.completed_at else req.created_at.isoformat(),
        })

    return {
        "total": total,
        "users": list(users_access.values())
    }


@router.get("/user-access/summary")
async def get_user_access_summary(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_superuser)
):
    """
    Сводная статистика по доступам пользователей.
    """
    # Общее количество пользователей с доступом
    users_with_access = db.query(func.count(func.distinct(AccessRequest.target_user_id))).filter(
        AccessRequest.status.in_([RequestStatus.APPROVED, RequestStatus.IMPLEMENTED])
    ).scalar()

    # Количество систем с выданными доступами
    systems_with_access = db.query(func.count(func.distinct(AccessRequest.system_id))).filter(
        AccessRequest.status.in_([RequestStatus.APPROVED, RequestStatus.IMPLEMENTED])
    ).scalar()

    # Всего активных доступов
    total_accesses = db.query(func.count(AccessRequest.id)).filter(
        AccessRequest.status.in_([RequestStatus.APPROVED, RequestStatus.IMPLEMENTED])
    ).scalar()

    # Временных доступов
    temporary_accesses = db.query(func.count(AccessRequest.id)).filter(
        and_(
            AccessRequest.status.in_([RequestStatus.APPROVED, RequestStatus.IMPLEMENTED]),
            AccessRequest.is_temporary == True
        )
    ).scalar()

    # Доступы по системам
    systems_stats = db.query(
        System.id,
        System.name,
        System.code,
        func.count(AccessRequest.id).label('access_count')
    ).join(
        AccessRequest, AccessRequest.system_id == System.id
    ).filter(
        AccessRequest.status.in_([RequestStatus.APPROVED, RequestStatus.IMPLEMENTED])
    ).group_by(System.id, System.name, System.code).order_by(
        func.count(AccessRequest.id).desc()
    ).limit(10).all()

    # Доступы по отделам
    dept_stats = db.query(
        User.department,
        func.count(AccessRequest.id).label('access_count')
    ).join(
        AccessRequest, AccessRequest.target_user_id == User.id
    ).filter(
        AccessRequest.status.in_([RequestStatus.APPROVED, RequestStatus.IMPLEMENTED]),
        User.department.isnot(None)
    ).group_by(User.department).order_by(
        func.count(AccessRequest.id).desc()
    ).limit(10).all()

    return {
        "users_with_access": users_with_access or 0,
        "systems_with_access": systems_with_access or 0,
        "total_accesses": total_accesses or 0,
        "temporary_accesses": temporary_accesses or 0,
        "by_system": [
            {"system_id": s.id, "system_name": s.name, "system_code": s.code, "count": s.access_count}
            for s in systems_stats
        ],
        "by_department": [
            {"department": d.department or "Не указан", "count": d.access_count}
            for d in dept_stats
        ]
    }


@router.get("/user-access/excel")
async def export_user_access_excel(
    system_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_superuser)
):
    """Экспорт отчёта по доступам в Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Получить все доступы
    query = db.query(AccessRequest).options(
        joinedload(AccessRequest.target_user),
        joinedload(AccessRequest.system),
        joinedload(AccessRequest.subsystem),
        joinedload(AccessRequest.access_role),
    ).filter(
        AccessRequest.status.in_([RequestStatus.APPROVED, RequestStatus.IMPLEMENTED])
    )

    if system_id:
        query = query.filter(AccessRequest.system_id == system_id)

    requests = query.order_by(
        AccessRequest.target_user_id,
        AccessRequest.system_id
    ).all()

    wb = Workbook()

    # Стили
    header_fill = PatternFill(start_color="16306C", end_color="16306C", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    data_font = Font(size=10)
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # ===== Лист: Доступы пользователей =====
    ws = wb.active
    ws.title = "Доступы пользователей"

    headers = [
        'Пользователь', 'Логин', 'Email', 'Отдел', 'Должность',
        'Система', 'Код системы', 'Подсистема', 'Роль доступа', 'Уровень доступа',
        'Статус', 'Временный', 'Действует с', 'Действует до', 'Дата выдачи'
    ]

    # Заголовки
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_alignment

    # Данные
    for row_idx, req in enumerate(requests, 2):
        data = [
            req.target_user.full_name if req.target_user else '—',
            req.target_user.username if req.target_user else '—',
            req.target_user.email if req.target_user else '—',
            req.target_user.department if req.target_user else '—',
            req.target_user.position if req.target_user else '—',
            req.system.name if req.system else '—',
            req.system.code if req.system else '—',
            req.subsystem.name if req.subsystem else '—',
            req.access_role.name if req.access_role else '—',
            req.access_role.access_level if req.access_role else '—',
            get_status_label(req.status.value if req.status else '—'),
            'Да' if req.is_temporary else 'Нет',
            format_date(req.valid_from) if req.valid_from else '—',
            format_date(req.valid_until) if req.valid_until else '—',
            format_date(req.completed_at) if req.completed_at else format_date(req.created_at),
        ]

        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = left_alignment

    # Ширина колонок
    column_widths = [25, 15, 30, 20, 20, 20, 12, 15, 18, 15, 12, 10, 14, 14, 16]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = 'A2'

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"user_access_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
